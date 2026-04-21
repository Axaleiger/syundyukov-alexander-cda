# -*- coding: utf-8 -*-
"""
data/actives.xlsx → data/actives_tree.json

Поддерживаются два формата:
1) Плоская таблица: A — № п/п, B — ДО, C — Месторождение, D — Пласт (пласты в JSON не попадают).
   Лист «Справочник» или первый лист; первая строка-шапка определяется по «ДО» / «Месторождение».
2) Старый формат: листы Actives, Fields, Layers.
"""
import hashlib
import json
import sys
from collections import defaultdict
from pathlib import Path
from typing import Optional

try:
    from openpyxl import load_workbook
except ImportError:
    print("pip install openpyxl", file=sys.stderr)
    sys.exit(1)

ROOT = Path(__file__).resolve().parents[1]
XLSX = ROOT / "data" / "actives.xlsx"
JSON_OUT = ROOT / "data" / "actives_tree.json"


def _s(v) -> str:
    if v is None:
        return ""
    if isinstance(v, float) and v == int(v):
        v = int(v)
    return str(v).strip()


def _aid(do_name: str) -> str:
    return "DO_" + hashlib.md5(do_name.encode("utf-8")).hexdigest()[:12]


def _fid(do_name: str, field_name: str, row_num: Optional[int] = None) -> str:
    """row_num — номер строки в Excel; если задан, одна и та же пара ДО+месторождение в разных строках даёт разные id."""
    key = f"{do_name}|{field_name}|r{row_num}" if row_num is not None else f"{do_name}|{field_name}"
    return "F_" + hashlib.md5(key.encode("utf-8")).hexdigest()[:14]


def _is_header_row(do_name: str, field_name: str) -> bool:
    """Строка-шапка: «ДО», «Месторождение» и т.п."""
    d, f = do_name.lower(), field_name.lower()
    if d in ("до", "д.о.") and ("местор" in f or f == "месторождение"):
        return True
    if do_name == "ДО" and field_name == "Месторождение":
        return True
    if "п/п" in do_name or do_name in ("№", "№ п/п", "N"):
        return True
    return False


def build_from_flat_sheet(ws):
    """Колонки A–D: № п/п, ДО, Месторождение, Пласт. В JSON — только ДО и месторождения, порядок как в листе.
    Каждая строка данных — отдельная позиция месторождения (дубликаты названия в одном ДО сохраняются)."""
    do_order: list[str] = []
    seen_do: set[str] = set()
    buckets: dict[str, dict] = {}
    excel_row = 0

    for row in ws.iter_rows(min_row=1, min_col=1, max_col=6, values_only=True):
        excel_row += 1
        if not row:
            continue
        cells = list(row) + [None] * 6
        _num, do_name, field_name, _layer = cells[0], cells[1], cells[2], cells[3]
        do_name = _s(do_name)
        field_name = _s(field_name)
        if _is_header_row(do_name, field_name):
            continue
        if not do_name:
            continue

        aid = _aid(do_name)
        if aid not in buckets:
            buckets[aid] = {"id": aid, "name": do_name, "_field_order": [], "_fields_map": {}}
            if do_name not in seen_do:
                seen_do.add(do_name)
                do_order.append(do_name)

        if not field_name:
            continue

        fid = _fid(do_name, field_name, excel_row)
        b = buckets[aid]
        b["_fields_map"][fid] = {"id": fid, "name": field_name}
        b["_field_order"].append(fid)

    if not buckets:
        return None

    actives = []
    for do_name in do_order:
        aid = _aid(do_name)
        a = buckets[aid]
        flds = [a["_fields_map"][fid] for fid in a["_field_order"]]
        actives.append({"id": a["id"], "name": a["name"], "fields": flds})

    # Одно и то же название месторождения в разных ДО → подсказка в UI
    name_to_aids: dict[str, list[str]] = defaultdict(list)
    for a in actives:
        for f in a["fields"]:
            name_to_aids[f["name"]].append(a["id"])

    field_shared_by_actives: dict[str, list[str]] = {}
    for a in actives:
        for f in a["fields"]:
            aids_same_name = sorted(set(name_to_aids[f["name"]]))
            if len(aids_same_name) > 1:
                field_shared_by_actives[f["id"]] = [x for x in aids_same_name if x != a["id"]]

    return {
        "meta": {
            "source": "actives.xlsx",
            "format": "flat_ABCD_fields_only",
            "columns": ["№ п/п", "ДО", "Месторождение", "Пласт"],
            "note": "В JSON только ДО и месторождения, порядок месторождений как в Excel. Колонка «Пласт» читается для пропуска строки-шапки; пласты в дерево не попадают. Пересборка: python scripts/actives_to_json.py",
        },
        "actives": actives,
        "field_shared_by_actives": field_shared_by_actives,
    }


def build_from_legacy(wb) -> dict:
    actives_rows = list(wb["Actives"].iter_rows(min_row=2, values_only=True))
    fields_rows = list(wb["Fields"].iter_rows(min_row=2, values_only=True))
    layers_rows = list(wb["Layers"].iter_rows(min_row=2, values_only=True))

    active_map = {}
    for r in actives_rows:
        if not r or not r[0]:
            continue
        aid, name = r[0], r[1] or ""
        active_map[str(aid)] = {"id": str(aid), "name": str(name), "fields": {}}

    for r in fields_rows:
        if not r or not r[0]:
            continue
        fid, fname, aid = str(r[0]), str(r[1] or ""), str(r[2] or "")
        if aid not in active_map:
            active_map[aid] = {"id": aid, "name": aid, "fields": {}}
        if fid not in active_map[aid]["fields"]:
            active_map[aid]["fields"][fid] = {"id": fid, "name": fname, "layers": []}

    for r in layers_rows:
        if not r or not r[0]:
            continue
        lid, lname, fid = str(r[0]), str(r[1] or ""), str(r[2] or "")
        for aid, a in active_map.items():
            if fid in a["fields"]:
                layers = a["fields"][fid]["layers"]
                if not any(x["id"] == lid for x in layers):
                    layers.append({"id": lid, "name": lname})

    actives = []
    for a in active_map.values():
        flds = list(a["fields"].values())
        flds.sort(key=lambda x: x["id"])
        actives.append({"id": a["id"], "name": a["name"], "fields": flds})
    actives.sort(key=lambda x: x["id"])

    field_to_actives = {}
    for r in fields_rows:
        if not r or not r[0]:
            continue
        fid, aid = str(r[0]), str(r[2] or "")
        field_to_actives.setdefault(fid, []).append(aid)

    return {
        "meta": {"source": "actives.xlsx", "format": "legacy_sheets"},
        "actives": actives,
        "field_shared_by_actives": {
            k: sorted(set(v)) for k, v in field_to_actives.items() if len(set(v)) > 1
        },
    }


def main():
    if not XLSX.is_file():
        print("Нет файла", XLSX, "- положите actives.xlsx в data/ или запустите create_actives_xlsx.py", file=sys.stderr)
        sys.exit(1)

    wb = load_workbook(XLSX, read_only=True, data_only=True)
    names = wb.sheetnames

    payload = None
    # Сначала ваш формат: лист «Справочник» или первая таблица A–D
    if "Справочник" in names:
        payload = build_from_flat_sheet(wb["Справочник"])
    elif "Actives" in names and "Fields" in names and "Layers" in names:
        payload = build_from_legacy(wb)
    else:
        payload = build_from_flat_sheet(wb[names[0]])

    if not payload or not payload.get("actives"):
        print("Не удалось разобрать таблицу: проверьте лист и столбцы A–D (ДО, Месторождение, Пласт).", file=sys.stderr)
        sys.exit(1)

    JSON_OUT.parent.mkdir(parents=True, exist_ok=True)
    JSON_OUT.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")
    n = len(payload["actives"])
    print(f"Записано: {JSON_OUT} — ДО: {n}, формат: {payload['meta'].get('format')}")


if __name__ == "__main__":
    main()
