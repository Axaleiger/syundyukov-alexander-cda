# -*- coding: utf-8 -*-
"""Сборка многостраничного Excel из data/asset_modeling_knowledge.json"""
from __future__ import annotations

import json
import sys
from pathlib import Path

try:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter
except ImportError:
    print("Установите openpyxl: pip install openpyxl", file=sys.stderr)
    sys.exit(1)

ROOT = Path(__file__).resolve().parents[1]
DATA_PATH = ROOT / "data" / "asset_modeling_knowledge.json"
INFLUENCE_PATH = ROOT / "data" / "influence_tags.json"
ACTIVES_TREE_PATH = ROOT / "data" / "actives_tree.json"
OUT_PATH = ROOT / "output" / "Asset_Modeling_Constructor.xlsx"

HEADER_FILL = PatternFill("solid", fgColor="1F4E79")
HEADER_FONT = Font(color="FFFFFF", bold=True)
WRAP = Alignment(wrap_text=True, vertical="top")


def autosize_columns(ws, max_width: int = 55):
    for col in range(1, ws.max_column + 1):
        letter = get_column_letter(col)
        best = 10
        for row in range(1, min(ws.max_row + 1, 200)):
            cell = ws.cell(row=row, column=col)
            if cell.value is None:
                continue
            ln = len(str(cell.value))
            best = min(max(best, ln + 2), max_width)
        ws.column_dimensions[letter].width = best


def add_sheet(wb: Workbook, name: str, headers: list[str], rows: list[list]):
    ws = wb.create_sheet(title=name[:31])
    for c, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=c, value=h)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = WRAP
    for r, row in enumerate(rows, 2):
        for c, val in enumerate(row, 1):
            cell = ws.cell(row=r, column=c, value=val)
            cell.alignment = WRAP
    autosize_columns(ws)
    return ws


def main():
    DATA_PATH.parent.mkdir(parents=True, exist_ok=True)
    if not DATA_PATH.is_file():
        print(f"Нет файла {DATA_PATH}", file=sys.stderr)
        sys.exit(1)

    with open(DATA_PATH, "r", encoding="utf-8") as f:
        data = json.load(f)

    OUT_PATH.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    wb.remove(wb.active)

    meta = data.get("meta", {})
    add_sheet(
        wb,
        "Meta",
        ["key", "value"],
        [
            ["version", meta.get("version", "")],
            ["locale", meta.get("locale", "")],
            ["description", meta.get("description", "")],
        ],
    )

    pr = data.get("scenario_principles", [])
    add_sheet(
        wb,
        "ScenarioPrinciples",
        ["id", "title", "text"],
        [[p["id"], p.get("title", ""), p.get("text", "")] for p in pr],
    )

    sm = data.get("scope_modes", [])
    add_sheet(
        wb,
        "ScopeModes",
        ["id", "code", "name", "description", "user_selects", "notes"],
        [
            [
                x["id"],
                x.get("code", ""),
                x.get("name", ""),
                x.get("description", ""),
                ", ".join(x.get("user_selects", [])),
                x.get("notes", ""),
            ]
            for x in sm
        ],
    )

    ar = data.get("allocation_rules", [])
    add_sheet(
        wb,
        "AllocationRules",
        ["id", "name", "description"],
        [[a["id"], a.get("name", ""), a.get("description", "")] for a in ar],
    )

    dim = data.get("dimensions", {})
    add_sheet(
        wb,
        "Dim_Objectives",
        ["id", "name", "direction", "semantics"],
        [
            [o["id"], o.get("name", ""), o.get("direction", ""), o.get("semantics", "")]
            for o in dim.get("objectives", [])
        ],
    )
    add_sheet(
        wb,
        "Dim_SuccessMetrics",
        ["id", "name"],
        [[m["id"], m.get("name", "")] for m in dim.get("success_metrics", [])],
    )
    add_sheet(
        wb,
        "Dim_Constraints",
        ["id", "name"],
        [[c["id"], c.get("name", "")] for c in dim.get("constraints", [])],
    )
    add_sheet(
        wb,
        "Dim_Levers",
        ["id", "name"],
        [[lev["id"], lev.get("name", "")] for lev in dim.get("levers", [])],
    )
    add_sheet(
        wb,
        "Dim_Horizons",
        ["id", "name"],
        [[t["id"], t.get("name", "")] for t in dim.get("horizons", [])],
    )
    add_sheet(
        wb,
        "Dim_ScenarioKinds",
        ["id", "name"],
        [[k["id"], k.get("name", "")] for k in dim.get("scenario_kinds", [])],
    )
    add_sheet(
        wb,
        "Dim_Bases",
        ["id", "name"],
        [[b["id"], b.get("name", "")] for b in dim.get("bases", [])],
    )

    kpi = data.get("kpi_codes", [])
    add_sheet(
        wb,
        "Dim_KPI",
        ["id", "name"],
        [[k["id"], k.get("name", "")] for k in kpi],
    )

    cat_rows = []
    drv_rows = []
    hyp_rows = []
    hyp_lev_rows = []

    for cat in data.get("fcf_bridge", {}).get("categories", []):
        cid = cat["id"]
        cat_rows.append([cid, cat.get("name", "")])
        for drv in cat.get("drivers", []):
            drv_rows.append(
                [
                    drv["id"],
                    cid,
                    drv.get("name", ""),
                    drv.get("typical_polarity", ""),
                    ", ".join(drv.get("kpi_impacts", [])),
                ]
            )
            for h in drv.get("hypotheses", []):
                hyp_rows.append(
                    [
                        h["id"],
                        drv["id"],
                        h.get("title", ""),
                        h.get("text", ""),
                        ", ".join(h.get("lever_ids", [])),
                    ]
                )
                for lid in h.get("lever_ids", []):
                    hyp_lev_rows.append([h["id"], lid])

    add_sheet(wb, "Influence_Categories", ["category_id", "name"], cat_rows)
    add_sheet(
        wb,
        "FCF_Drivers",
        ["driver_id", "category_id", "name", "typical_polarity", "kpi_impacts"],
        drv_rows,
    )
    add_sheet(
        wb,
        "Hypotheses",
        ["hypothesis_id", "driver_id", "title", "text", "lever_ids_concat"],
        hyp_rows,
    )
    add_sheet(
        wb,
        "Hypothesis_Lever",
        ["hypothesis_id", "lever_id"],
        hyp_lev_rows,
    )

    hints = data.get("lever_constraint_hints", [])
    add_sheet(
        wb,
        "Lever_Constraint_Hints",
        ["lever_id", "constraint_id", "compatible", "note"],
        [
            [h["lever_id"], h["constraint_id"], h.get("compatible", ""), h.get("note", "")]
            for h in hints
        ],
    )

    pt = data.get("prompt_templates", [])
    add_sheet(
        wb,
        "PromptTemplates",
        ["template_id", "name", "template"],
        [[p["id"], p.get("name", ""), p.get("template", "")] for p in pt],
    )

    dh = data.get("dimension_help", {})
    help_rows: list[list] = []
    for section, items in sorted(dh.items()):
        for iid, text in sorted((items or {}).items()):
            help_rows.append([section, iid, text])
    add_sheet(wb, "DimensionHelp", ["section", "id", "help_text"], help_rows)

    if INFLUENCE_PATH.is_file():
        with open(INFLUENCE_PATH, "r", encoding="utf-8") as f:
            inf = json.load(f)
        o_rows = []
        for oid, tags in (inf.get("objective_tag_activation") or {}).items():
            for t in tags:
                o_rows.append([oid, t])
        add_sheet(wb, "Tags_by_Objective", ["objective_id", "tag"], sorted(o_rows))

        c_rows = []
        for cid, tags in (inf.get("constraint_tag_activation") or {}).items():
            for t in tags:
                c_rows.append([cid, t])
        add_sheet(wb, "Tags_by_Constraint", ["constraint_id", "tag"], sorted(c_rows))

        k_rows = []
        for kid, tags in (inf.get("kind_tag_activation") or {}).items():
            for t in tags:
                k_rows.append([kid, t])
        add_sheet(wb, "Tags_by_ScenarioKind", ["scenario_kind_id", "tag"], sorted(k_rows))

        cd_rows = [[cid, ", ".join(tags)] for cid, tags in sorted((inf.get("category_default_tags") or {}).items())]
        add_sheet(wb, "Category_Default_Tags", ["category_id", "tags_csv"], cd_rows)

        ho_rows = [[hid, ", ".join(tags)] for hid, tags in sorted((inf.get("hypothesis_tag_overrides") or {}).items())]
        add_sheet(wb, "Hypothesis_Tag_Overrides", ["hypothesis_id", "tags_csv"], ho_rows)

    add_sheet(
        wb,
        "Actives_Data_Hint",
        ["instruction"],
        [
            [
                "Число строк в книге велико из‑за листов знаний (гипотезы, драйверы FCF, справки по измерениям и т.д.) — это НЕ количество ДО. "
                "Справочник активов (ДО и месторождения) — на листе «Справочник_ДО» (по одной строке на месторождение; колонки пласта/layer_id могут быть пустыми). "
                "Исходник: data/actives.xlsx, лист «Справочник», колонки A–D (№, ДО, Месторождение, Пласт) → python scripts/actives_to_json.py → actives_tree.json."
            ]
        ],
    )

    if ACTIVES_TREE_PATH.is_file():
        with open(ACTIVES_TREE_PATH, "r", encoding="utf-8") as f:
            atree = json.load(f)
        spr_rows: list[list] = []
        n = 0
        for act in atree.get("actives", []):
            do_name = act.get("name", "")
            aid = act.get("id", "")
            for fld in act.get("fields", []):
                fn = fld.get("name", "")
                fid = fld.get("id", "")
                layers = fld.get("layers") or []
                if not layers:
                    n += 1
                    spr_rows.append([n, do_name, fn, "", aid, fid, ""])
                else:
                    for lyr in layers:
                        n += 1
                        spr_rows.append(
                            [
                                n,
                                do_name,
                                fn,
                                lyr.get("name", ""),
                                aid,
                                fid,
                                lyr.get("id", ""),
                            ]
                        )
        add_sheet(
            wb,
            "Справочник_ДО",
            ["№ п/п", "ДО", "Месторождение", "Пласт", "asset_id", "field_id", "layer_id"],
            spr_rows,
        )
    else:
        add_sheet(
            wb,
            "Справочник_ДО",
            ["№ п/п", "ДО", "Месторождение", "Пласт", "asset_id", "field_id", "layer_id"],
            [],
        )

    # Пустые листы-заготовки для экземпляров (заполняются вручную или из платформы)
    add_sheet(
        wb,
        "Instance_Assets",
        ["asset_id", "name", "comment"],
        [["A-001", "Пример: ДО / актив", "Замените на свои ID"]],
    )
    add_sheet(
        wb,
        "Instance_Fields",
        ["field_id", "name", "comment"],
        [["F-001", "Пример: месторождение", ""]],
    )
    add_sheet(
        wb,
        "Rel_Asset_Field",
        ["asset_id", "field_id", "share_pct_optional", "role"],
        [
            ["A-001", "F-001", "100", "operator"],
            ["A-002", "F-001", "40", "partner — пример мульти-актива"],
        ],
    )
    add_sheet(
        wb,
        "User_Prompt_Log",
        ["timestamp", "assembled_prompt", "scope_mode", "entity_note"],
        [],
    )

    wb.save(OUT_PATH)
    print(f"Записано: {OUT_PATH}")


if __name__ == "__main__":
    main()
