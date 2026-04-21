# -*- coding: utf-8 -*-
"""
Импорт actives_(only_fields).xlsx → data/actives.xlsx (лист «Справочник»).

Формат исходника: строка 1 — шапка; A — № (в выход попадает для порядка); B — ДО; C — месторождение.
Далее: python scripts/actives_to_json.py (вызывается автоматически).
"""
import subprocess
import sys
from pathlib import Path

try:
    from openpyxl import Workbook, load_workbook
except ImportError:
    print("pip install openpyxl", file=sys.stderr)
    sys.exit(1)

ROOT = Path(__file__).resolve().parents[1]
SRC = ROOT / "actives_(only_fields).xlsx"
DST = ROOT / "data" / "actives.xlsx"


def _s(v):
    if v is None:
        return ""
    if isinstance(v, float) and v == int(v):
        v = int(v)
    return str(v).strip()


def main():
    if not SRC.is_file():
        print("Нет файла:", SRC, file=sys.stderr)
        sys.exit(1)

    wb_src = load_workbook(SRC, read_only=True, data_only=True)
    ws = wb_src.active
    rows_out = []
    for row_idx, row in enumerate(ws.iter_rows(min_row=1, values_only=True), start=1):
        if not row:
            continue
        cells = list(row) + [None] * 4
        num, do_name, field_name = cells[0], cells[1], cells[2]
        if row_idx == 1:
            continue
        do_name = _s(do_name)
        field_name = _s(field_name)
        if not do_name and not field_name:
            continue
        if num is None:
            num = len(rows_out) + 1
        rows_out.append((num, do_name, field_name, ""))
    wb_src.close()

    wb = Workbook()
    ws2 = wb.active
    ws2.title = "Справочник"
    ws2.append(["№ п/п", "ДО", "Месторождение", "Пласт"])
    for r in rows_out:
        ws2.append(list(r))

    DST.parent.mkdir(parents=True, exist_ok=True)
    wb.save(DST)
    print("Импортировано строк данных:", len(rows_out), "->", DST)

    r = subprocess.run(
        [sys.executable, str(ROOT / "scripts" / "actives_to_json.py")],
        cwd=str(ROOT),
    )
    sys.exit(r.returncode)


if __name__ == "__main__":
    main()
