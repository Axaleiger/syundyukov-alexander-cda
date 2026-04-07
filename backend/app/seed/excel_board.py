"""
Парсинг Excel-доски по контракту frontend/src/modules/planning/data/bpmExcel.js
"""

from __future__ import annotations

from dataclasses import dataclass
from datetime import date, datetime, timedelta
from typing import Any

from openpyxl import load_workbook

EXCEL_EPOCH = date(1899, 12, 30)


def _serial_to_date(serial: Any) -> date | None:
    if serial is None or serial == "":
        return None
    try:
        n = float(serial)
    except (TypeError, ValueError):
        return None
    if n != n:  # NaN
        return None
    return EXCEL_EPOCH + timedelta(days=int(n))


REQUIRED_COLUMNS = [
    "Этап Название",
    "Карточка ID",
    "Карточка Название",
    "Исполнитель",
    "Согласующий",
    "Срок сдачи",
    "Статус",
    "Дата создания",
    "Используемые системы",
    "Входные данные",
    "Выходные данные",
]


@dataclass
class BoardParseResult:
    stages: list[str]
    # stage_name -> list of cards (card_id -> card dict with entries)
    tasks: dict[str, list[dict[str, Any]]]
    connections: list[dict[str, str]]


def parse_board_from_excel(path: str) -> BoardParseResult:
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    rows_iter = ws.iter_rows(values_only=True)
    header = next(rows_iter, None)
    if not header:
        return BoardParseResult(stages=[], tasks={}, connections=[])

    cols = [str(c).strip() if c is not None else "" for c in header]
    col_index = {name: i for i, name in enumerate(cols)}

    def cell(row: tuple[Any, ...], name: str) -> Any:
        i = col_index.get(name)
        if i is None or i >= len(row):
            return None
        return row[i]

    missing = [c for c in REQUIRED_COLUMNS if c not in col_index]
    if missing:
        wb.close()
        raise ValueError(f"Не хватает колонок: {missing}")

    stages: list[str] = []
    tasks: dict[str, dict[str, dict[str, Any]]] = {}
    seen_stages: set[str] = set()

    for row in rows_iter:
        if not row or all(x is None or str(x).strip() == "" for x in row):
            continue
        stage = cell(row, "Этап Название")
        if stage is None or str(stage).strip() == "":
            continue
        stage = str(stage).strip()
        if stage not in seen_stages:
            seen_stages.add(stage)
            stages.append(stage)
            tasks[stage] = {}

        card_id = cell(row, "Карточка ID")
        if card_id is None or str(card_id).strip() == "":
            continue
        card_id = str(card_id).strip()

        bucket = tasks[stage]
        if card_id not in bucket:
            dl = cell(row, "Срок сдачи")
            deadline = _serial_to_date(dl)
            if deadline is None and isinstance(dl, datetime):
                deadline = dl.date()
            dct_raw = cell(row, "Дата создания")
            date_str = ""
            if isinstance(dct_raw, datetime):
                date_str = dct_raw.strftime("%d.%m.%Y")
            elif dct_raw is not None:
                date_str = str(dct_raw).strip()

            bucket[card_id] = {
                "id": card_id,
                "name": str(cell(row, "Карточка Название") or "").strip(),
                "executor": str(cell(row, "Исполнитель") or "").strip(),
                "approver": str(cell(row, "Согласующий") or "").strip(),
                "deadline": deadline,
                "status": str(cell(row, "Статус") or "в работе").strip(),
                "date": date_str,
                "entries": [],
            }

        system = cell(row, "Используемые системы")
        system = str(system).strip() if system is not None else ""
        if system == "":
            continue
        inp = str(cell(row, "Входные данные") or "").strip()
        out = str(cell(row, "Выходные данные") or "").strip()
        bucket[card_id]["entries"].append({"system": system, "input": inp, "output": out})

    wb.close()

    out_tasks: dict[str, list[dict[str, Any]]] = {s: list(tasks[s].values()) for s in stages}
    conns = parse_connections_from_excel(path)
    return BoardParseResult(stages=stages, tasks=out_tasks, connections=conns)


def parse_connections_from_excel(path: str) -> list[dict[str, str]]:
    try:
        wb = load_workbook(path, read_only=True, data_only=True)
    except Exception:
        return []
    try:
        if "Связи" in wb.sheetnames:
            ws = wb["Связи"]
            rows = list(ws.iter_rows(values_only=True))
            wb.close()
            out: list[dict[str, str]] = []
            for row in rows[1:]:
                if not row or len(row) < 4:
                    continue
                out.append(
                    {
                        "fromStage": str(row[0] or "").strip(),
                        "fromId": str(row[1] or "").strip(),
                        "toStage": str(row[2] or "").strip(),
                        "toId": str(row[3] or "").strip(),
                    }
                )
            return [c for c in out if c["fromStage"] and c["fromId"] and c["toStage"] and c["toId"]]

        ws = wb[wb.sheetnames[0]]
        aoa = list(ws.iter_rows(values_only=True))
        wb.close()
        conn_row_idx = -1
        for i, row in enumerate(aoa):
            if row and str(row[0]).strip() == "Связи":
                conn_row_idx = i
                break
        if conn_row_idx < 0:
            return []
        header_row = aoa[conn_row_idx + 1] if conn_row_idx + 1 < len(aoa) else None
        if not header_row or str(header_row[0]).strip() != "Этап от":
            return []
        out = []
        for row in aoa[conn_row_idx + 2 :]:
            if not row:
                continue
            if not any(row[:4]):
                continue
            out.append(
                {
                    "fromStage": str(row[0] or "").strip(),
                    "fromId": str(row[1] or "").strip(),
                    "toStage": str(row[2] or "").strip(),
                    "toId": str(row[3] or "").strip(),
                }
            )
        return [c for c in out if c["fromStage"] and c["fromId"] and c["toStage"] and c["toId"]]
    except Exception:
        try:
            wb.close()
        except Exception:
            pass
        return []
